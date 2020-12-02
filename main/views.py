from collections import defaultdict

from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import OuterRef, Q
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import redirect, render, reverse, get_object_or_404
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.comments import Comment

from main.forms import AddTableForm
from main.models import *
from users.models import Person
from reports.models import *
# Create your views here.


def index(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse('users:login'))
    else:
        return HttpResponseRedirect(reverse('tables:start'))


@login_required(login_url='/users/login/')
def start(request):
    return render(request, "main/main.html")


@login_required(login_url='/users/login/')
def admins(request):
    return render(request, "main/admin.html")


@login_required(login_url='/users/login/')
def indicators(request):
    tables = Table.objects.all()
    if request.user.groups.filter(name='Operator').exists():
        tables = tables.exclude(tag__isnull=True)
    dict_tables = defaultdict(list)
    for table in tables:
        dict_tables[table.tag.get_value() if table.tag is not None else 'Справочник'].append(table)
    tags = list(Cell.objects.filter(row__table__name__icontains='Группы показателей', col__brief_name__icontains='Название').values_list('value__char200_value', flat=True))
    return render(request, 'main/tables_list_copy.html', { 'tags': tags, 'dict_tables': dict_tables })


@login_required(login_url='/users/login/')
def tables_list(request):
    add_form = AddTableForm()
    tables = Table.objects.all()
    if request.user.groups.filter(name='Operator').exists():
        tables = tables.exclude(tag__isnull=True)
    return render(request, 'main/tables_list.html', { 'form': add_form, 'tables': tables })


@login_required(login_url='/users/login/')
@permission_required('main.view_cell', raise_exception=True)
def table_data(request, table_id):
    table = get_object_or_404(Table, pk=table_id)
#    data, rows_number = get_table_cells(table_id, request.user)
    refs = get_table_refs(table_id)
    limits, field_id = get_limits(table_id)
    reports = Report.objects.filter(Q(state=EDITING)|Q(state=REWORK)|Q(state=NEW))
    if request.user.groups.filter(name='Operator').exists():
        operator = request.user.person_set.get()
        if operator.is_regional:
            reports = reports.filter(user__person__villages__in=operator.villages.values_list('id'))
        else:
            reports = reports.filter(user=request.user)
    return render(request, 'main/table_details.html', {'table': table, 'refs': refs, 'limits': limits, 'reports': reports, 'field_id': field_id})


@login_required(login_url='/users/login/')
@permission_required('main.view_column', raise_exception=True)
def table_fields(request, table_id):
    table = get_object_or_404(Table, pk=table_id)
    tables = Table.objects.all()
    columns = table.fields.select_related('table', 'parent').all()
    return render(request, 'main/table_fields.html', {'columns': columns, 'types': TYPES, 'tables': tables, 'table': table})


def get_table_cells(table_id, user, page=0, page_size=0):
    result = []
    rows = Row.objects.filter(table__id=table_id).values_list('id', flat=True)
    if user.groups.filter(name='Operator').exists():
        oper = user.person_set.get()
        rows = Cell.objects.filter(row__table__id=table_id, value__ref_value__id__in=oper.villages.values_list('id', flat=True)).values_list('row__id', flat=True)
    number_of_rows = len(rows)
    if page > 0 and page_size > 0:
        start = page_size * (page - 1)
        end = page_size * page
        rows = rows[start:end]
    all_cells = Cell.objects.select_related('row', 'col', 'row__report').filter(row__id__in=rows).order_by("row__number")
    row_data = {'id': -1}
    for cell in all_cells:
        if cell.row.id != row_data['id']:
            if row_data['id'] >= 0:
                result.append(row_data)
            row_data = {'id': cell.row.id,
                        'number': cell.row.number,
                        'date': cell.row.report.id if cell.row.report is not None else None}
        row_data[cell.col.id] = cell.get_value()
        if cell.col.column_type == REFERENCE and row_data[cell.col.id] is not None:
            row_data[cell.col.id] = row_data[cell.col.id].id
    if row_data['id'] >= 0:
        result.append(row_data)
    return result, number_of_rows


def get_table_refs(table_id):
    result = []
    try:
        table = Table.objects.select_related('tag').get(pk=table_id)
        fields = Column.objects.select_related('parent').filter(table__id=table_id, column_type=REFERENCE)
        for field in fields:
            col = field.parent.fields.filter(brief_name__icontains='Название').first()
            data = []
            cells = Cell.objects.select_related('row').filter(row__table__id=field.parent.id, col__id=col.id).order_by('row__number')
            if table.tag is not None and field.brief_name.lower() == 'показатель':
                row_ids = Cell.objects.filter(value__ref_value__id=table.tag.id).values_list('row__id', flat=True)
                cells = cells.filter(row__id__in=row_ids)
            for cell in cells:
                value = str(cell.get_value())
                additional_value = '('
                for additional_cell in Cell.objects.filter(row__id=cell.row.id, col__use_in_relation=True):
                    additional_value += str(additional_cell.get_value().get_value()) if additional_cell is not None and additional_cell.get_value() is not None else ''
                additional_value += ')'
                if additional_value != '()':
                    value = value + ' ' + additional_value
                data.append({'id': cell.id, 'value': value if value is not None else ''})
            result.append({'id': field.id, 'data': data})
    except Exception as e:
        result = []
    return result


def get_limits(table_id):
    result = {}
    table = Table.objects.select_related('tag').get(pk=table_id)
    field_id = None
    if table.tag is not None:
        field = Column.objects.select_related('parent').filter(table__id=table_id,
                                                               brief_name__icontains='Показатель').first()
        if field is not None:
            sq_max = Cell.objects.filter(col__brief_name__icontains='Макс. значение',
                                     row__id=OuterRef('row__id')).values('value__float_value')
            sq_min = Cell.objects.filter(col__brief_name__icontains='Мин. значение',
                                     row__id=OuterRef('row__id')).values('value__float_value')
            row_ids = Cell.objects.filter(value__ref_value__id=table.tag.id).values_list('row__id', flat=True)
            field_id = field.id
            cells = Cell.objects.filter(col__brief_name='Название', row__table__id=field.parent.id,
                                        row__id__in=row_ids).annotate(max=sq_max, min=sq_min).values_list('id', 'max', 'min')
            for cell in cells:
                result[cell[0]] = {'max': cell[1], 'min': cell[2]}
    return result, field_id


def export_indicators(request):
    wb = get_indicators()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=indicators.xlsx'
    wb.save(response)
    return response


def get_indicators():
    wb = Workbook()
    ws = wb.active
    indicators = Cell.objects.select_related('row', 'col', 'row__table', 'col__parent').filter(row__table__id=22).order_by('row__id', 'col__number')
    groups = dict(Cell.objects.filter(row__table__id=20).values_list('id', 'value__char200_value'))
    group = indicators[0].value.ref_value.id
    row_id = indicators[0].row.id
    ws.title = groups[group][:30]
    row_num = add_headers(ws)
    first_row = row_num
    group_field_id = indicators[0].col.id
    min_val = 0
    red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')

    for cell in indicators:
        if cell is None:
            continue
        if cell.col.parent is not None and cell.col.parent.id == 20 and group != cell.value.ref_value.id:
            ws.cell(row=row_num, column=1, value=(row_num - first_row + 1))
            group = cell.value.ref_value.id
            ws = wb.create_sheet(groups[group][:30])
            row_num = add_headers(ws)
            row_id = cell.row.id
            first_row = row_num

        if cell.row.id > row_id:
            ws.cell(row=row_num, column=1, value=(row_num - first_row + 1))
            row_id = cell.row.id
            row_num += 1
        if cell.col.id == group_field_id:
            continue
        elif cell.col.id == 53:
            min_val = cell.get_value()
        elif cell.col.id == 54:
            max_val = cell.get_value()
            current_cell = ws.cell(row=row_num, column=4)
            if min_val is not None:
                ws.conditional_formatting.add(current_cell.coordinate, CellIsRule(operator='lessThan', formula=[min_val], stopIfTrue=True, fill=red_fill))
            if max_val is not None:
                ws.conditional_formatting.add(current_cell.coordinate, CellIsRule(operator='greaterThan', formula=[max_val], stopIfTrue=True, fill=red_fill))
            add_comment(current_cell, min_val, max_val)
        elif cell.value is not None:
            if cell.col.column_type == REFERENCE:
                if cell.value.ref_value is not None:
                    ws.cell(row=row_num, column=(cell.col.number), value=cell.value.ref_value.get_value())
            else:
                c = ws.cell(row=row_num, column=(cell.col.number), value=cell.get_value())
                c.alignment = Alignment(wrapText=True)
    return wb


def add_headers(worksheet):
    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='Наименование показателя')
    worksheet.cell(row=1, column=3, value='Ед. изм.')
    worksheet.cell(row=1, column=4, value='Значение')
    worksheet.column_dimensions['B'].width = 80
    worksheet.column_dimensions['A'].width = 4
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    return 2


def add_conditional_formatting(worksheet, row, col, max_val, min_val):
    pass


def add_comment(cell, min_val, max_val):
    comment = ''
    if max_val is not None and min_val is not None:
        comment = 'Значение показателя должно быть между ' + str(min_val) + ' и ' + str(max_val)
    elif max_val is not None and min_val is None:
        comment = 'Значение показателя не должно превышать ' + str(max_val)
    elif min_val is not None and max_val is None:
        comment = 'Значение показателя не должно быть меньше ' + str(min_val)
    if comment != '':
        cell.comment = Comment(comment, '')