from collections import defaultdict

from django.shortcuts import get_object_or_404
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill, Font, NamedStyle
from openpyxl import Workbook
from openpyxl.comments import Comment

from main.forms import AddTableForm
from main.models import *
from users.models import Person
from reports.models import *


def export_table_to_excel(table_id):
    row = 1
    t = get_object_or_404(Table, pk=table_id)
    fields = t.all_fields
    data = Cell.objects.select_related('row', 'col', 'row__report').filter(row__table__id=table_id).order_by('row__number', 'col__number')
    wb = Workbook()
    ws = wb.active
    row = add_headers(ws, row, fields)
    add_data(ws, row, data, fields)
    return wb


def add_headers(worksheet, row, fields):
    c = worksheet.cell(row=row, column=1, value='№')
    c.font = Font(bold=True)
    c = worksheet.cell(row=row, column=2, value='Период')
    c.font = Font(bold=True)
    col = 3
    for field in fields:
        c = worksheet.cell(row=row, column=col, value=field.brief_name)
        c.font = Font(bold=True)
        col += 1
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['A'].width = 4
    return row + 1


def add_data(worksheet, row, data, fields):
    row_number = -1
    col = 1
    row -= 1
    for cell in data:
        if row_number != cell.row.number:
            row_number = cell.row.number
            row += 1
            worksheet.cell(row=row, column=1, value=row)
            if cell.row.report is not None:
                c = worksheet.cell(row=row, column=2, value=cell.row.report.report_date)
                c.number_format = 'DD.MM.YYYY'
            col = 3
        if cell.value is not None and cell.col.column_type == REFERENCE and cell.get_value() is not None:
            worksheet.cell(row=row, column=col, value=cell.get_value().get_value())
        elif cell.value is not None and cell.col.column_type != REFERENCE:
            c = worksheet.cell(row=row, column=col, value=cell.get_value())
            if cell.col.column_type == DATE or cell.col.column_type == DATETIME:
                c.number_format = 'DD/MM/YYYY'
        col += 1
    return row
