import datetime
import os

from django.conf import settings
from django.db import transaction
from openpyxl import Workbook, load_workbook
from imports.create_obj import ImportObjects
from imports.models import *


def import_excel(template_id, filename, user):
    if user.is_superuser or user.groups.filter(name='Manager').exists():
        villages = list(Cell.objects.filter(row__table__id=14, col__brief_name='Название').values_list('id', flat=True))
    else:
        villages = list(user.person_set.get().values_list('id', flat=True))
    import_obj = ImportObjects(villages)
    temp = ListTemplate.objects.get(pk=template_id)
    wb = load_workbook(os.path.join(settings.BASE_DIR, 'uploads', filename))
    wb_res = ''
    try:
        with transaction.atomic():
            for sheetname in wb.sheetnames:
                wb_res += import_sheet(temp, import_obj, wb[sheetname])
            import_obj.create_all()
    except Exception as e:
        wb_res += str(e)
    return build_results(temp, filename, user, wb_res)


def get_error_message(error_no, indicator, oktmo, value, row):
    if error_no == 1:
        return 'Значение ' + str(value) + ' показателя ' + indicator + ' для наслега с ОКТМО ' + oktmo + ' не удволетворяет ограничениям! В строке ' + str(row) + '\n'
    elif error_no == 2:
        return ' ' + indicator + ' для наслега  с ОКТМО ' + oktmo + ' ' + str(value) + ' ! В строке ' + str(row) + '\n'
    elif error_no == 3:
        return 'Значение ' + str(value) + ' показателя ' + indicator + ' для наслега  с ОКТМО ' + oktmo + ' не является числом! В строке ' + str(row) + '\n'
    elif error_no == 4:
        return 'Данные для наслега  с ОКТМО ' + oktmo + ' вы не можете загружать данные!  В строке ' + str(row) + '\n'
    elif error_no == 5:
        return 'Индикатор "' + indicator + '" не представлен в системе!\n'
    return ''


def import_sheet(import_template, import_obj, ws):
    row = 2
    sheet_res = ''
    while True:
        year = ws.cell(row=row, column=import_template.year_col).value
        oktmo = ws.cell(row=row, column=import_template.oktmo_col).value
        indicator = ws.cell(row=row, column=import_template.indicator_col).value
        value = ws.cell(row=row, column=import_template.value_col).value
        if year is None and oktmo is None and indicator is None and value is None:
            break
        report_date = datetime.datetime.strptime('01.01.' + str(int(year) + 1), "%d.%m.%Y")
        oktmo = str(oktmo)
        res = import_obj.add_object(report_date, oktmo, indicator, value)
        sheet_res += get_error_message(res, indicator, oktmo, value, row)
        row += 1
    return sheet_res


def build_results(template, filename, user, results):
    result = ImportResult()
    result.template = template
    result.upload_user = user
    result.file_name = filename
    result.result = results
    result.save()
    return result